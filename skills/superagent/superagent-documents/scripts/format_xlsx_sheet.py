"""
Per-sheet XLSX formatter — makes any data sheet look professional.

Usage:
    from format_xlsx_sheet import format_data_sheet, force_widths_xml

    format_data_sheet(
        ws,
        title="PENDAPATAN",
        col_widths={'A': 13, 'B': 18, 'C': 30, 'D': 12, 'E': 18, 'F': 10},
        data_start_row=3,
        last_data_row=22,
        last_col_letter='F',
        total_row=23,
        date_cols=('A',),
        money_cols=('E',),
        has_status=True,
    )

    wb.save('/path/to/file.xlsx')
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile, shutil, re
import subprocess, os

# Color palette — warm "Café Latte" theme
C_HEADER_BG = '5D4037'    # dark chocolate
C_HEADER_FG = 'FFFFFF'
C_TITLE_BG  = '8B6F47'    # warm brown
C_TITLE_FG  = 'FFFFFF'
C_ROW_ODD   = 'F5E6D3'    # cream
C_ROW_EVEN  = 'FFFAF3'    # off-white
C_TOTAL_BG  = '2C5F2D'    # dark green
C_TOTAL_FG  = 'FFFFFF'
C_BORDER    = '8B6F47'
C_ACTIVE_BG = 'E8F5E9'
C_ARCHIVED_BG = 'FFEBEE'
C_DANGER_BG = 'FFCDD2'

thin = Border(
    left=Side(style='thin', color=C_BORDER),
    right=Side(style='thin', color=C_BORDER),
    top=Side(style='thin', color=C_BORDER),
    bottom=Side(style='thin', color=C_BORDER)
)
thick = Border(
    left=Side(style='medium', color=C_TOTAL_BG),
    right=Side(style='medium', color=C_TOTAL_BG),
    top=Side(style='medium', color=C_TOTAL_BG),
    bottom=Side(style='medium', color=C_TOTAL_BG)
)


def format_data_sheet(ws, *, col_widths, data_start_row, last_data_row,
                     last_col_letter, total_row=None, freeze_row=3,
                     date_cols=(), money_cols=(), int_cols=(), has_status=False):
    """Apply borders, alternating colors, alignment, status badges, and total row."""
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    last_col_idx = ord(last_col_letter) - ord('A') + 1

    # Title row — merged banner if cell has content
    if ws.cell(row=1, column=1).value:
        try:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_idx)
        except Exception:
            pass  # already merged
        t = ws.cell(row=1, column=1)
        t.font = Font(bold=True, size=16, color=C_TITLE_FG)
        t.fill = PatternFill('solid', fgColor=C_TITLE_BG)
        t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header row (one above data)
    hdr_row = data_start_row - 1
    for c in range(1, last_col_idx + 1):
        h = ws.cell(row=hdr_row, column=c)
        h.font = Font(bold=True, size=11, color=C_HEADER_FG)
        h.fill = PatternFill('solid', fgColor=C_HEADER_BG)
        h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        h.border = thin
    ws.row_dimensions[hdr_row].height = 24

    # Data rows
    for r in range(data_start_row, last_data_row + 1):
        for c in range(1, last_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            col = get_column_letter(c)
            cell.fill = PatternFill('solid',
                fgColor=C_ROW_ODD if (r - data_start_row) % 2 == 0 else C_ROW_EVEN)
            if col in date_cols:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = 'DD/MM/YYYY'
            elif col in money_cols:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '"Rp"#,##0'
            elif col in int_cols:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if has_status and col == last_col_letter:
                v = str(cell.value or '').lower()
                if 'active' in v:
                    cell.fill = PatternFill('solid', fgColor=C_ACTIVE_BG)
                    cell.font = Font(bold=True, color=C_TOTAL_BG, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif 'archived' in v or 'deleted' in v:
                    cell.fill = PatternFill('solid', fgColor=C_ARCHIVED_BG)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Total row
    if total_row and total_row <= last_data_row + 1:
        for c in range(1, last_col_idx + 1):
            t = ws.cell(row=total_row, column=c)
            t.font = Font(bold=True, size=11, color=C_TOTAL_FG)
            t.fill = PatternFill('solid', fgColor=C_TOTAL_BG)
            t.border = thick
            col = get_column_letter(c)
            if col in money_cols:
                t.alignment = Alignment(horizontal='right', vertical='center')
                t.number_format = '"Rp"#,##0'
            elif col in int_cols:
                t.alignment = Alignment(horizontal='right', vertical='center')
                t.number_format = '#,##0'
            else:
                t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[total_row].height = 24

    if freeze_row:
        ws.freeze_panes = f'A{freeze_row}'


def trim_empty_rows(ws, keep_buffer=5):
    """Delete empty trailing rows safely (handles merged ranges)."""
    last_with_data = 1
    for r in range(1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None
               for c in range(1, ws.max_column + 1)):
            last_with_data = r
    target_end = last_with_data + keep_buffer
    for mr in list(ws.merged_cells.ranges):
        if mr.max_row > target_end:
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass
    if ws.max_row > target_end:
        ws.delete_rows(target_end + 1, ws.max_row - target_end)


def recalc_via_libreoffice(path):
    """Round-trip through ODS so formulas get cached values."""
    tmp = '/tmp/recalc_' + os.path.basename(path).replace('.xlsx', '')
    shutil.rmtree(tmp, ignore_errors=True)
    os.makedirs(tmp, exist_ok=True)
    subprocess.run(['soffice', '--headless', '--calc', '--convert-to', 'ods',
                    '--outdir', tmp, path], capture_output=True, timeout=120)
    ods_file = tmp + '/' + os.path.basename(path).replace('.xlsx', '.ods')
    out_dir = tmp + '/out'
    os.makedirs(out_dir, exist_ok=True)
    subprocess.run(['soffice', '--headless', '--calc', '--convert-to', 'xlsx',
                    '--outdir', out_dir, ods_file], capture_output=True, timeout=120)
    shutil.copy(out_dir + '/' + os.path.basename(ods_file).replace('.ods', '.xlsx'), path)


def force_widths_xml(path):
    """Post-process xlsx XML to ensure Excel respects openpyxl column widths
    (bypasses LibreOffice recalc reset)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    desired = {}
    for sn in wb.sheetnames:
        desired[sn] = {}
        for col_letter, dim in wb[sn].column_dimensions.items():
            if dim.width:
                desired[sn][col_letter] = dim.width

    with zipfile.ZipFile(path, 'r') as z:
        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

    name_to_rid = dict(re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml))
    rid_to_target = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    sheet_to_internal = {n: 'xl/' + rid_to_target[name_to_rid[n]].lstrip('/')
                         for n in name_to_rid}

    tmp = '/tmp/force_widths.tmp.xlsx'
    with zipfile.ZipFile(path, 'r') as zin, \
         zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            for sheet_name, widths in desired.items():
                if item == sheet_to_internal.get(sheet_name):
                    xml = data.decode('utf-8')
                    xml = re.sub(r'<cols>.*?</cols>', '', xml, flags=re.DOTALL)
                    cols_xml = '<cols>'
                    for col_letter, width in widths.items():
                        col_idx = ord(col_letter) - ord('A') + 1
                        cols_xml += f'<col customWidth="1" min="{col_idx}" max="{col_idx}" width="{width}"/>'
                    cols_xml += '</cols>'
                    xml = xml.replace('<sheetData', cols_xml + '<sheetData', 1)
                    data = xml.encode('utf-8')
                    break
            zout.writestr(item, data)
    shutil.copy(tmp, path)


def find_last_data_row(ws, start=3, col=1):
    """Find the last row with data in the given column."""
    last = start - 1
    for r in range(start, ws.max_row + 1):
        if ws.cell(row=r, column=col).value is not None:
            last = r
    return last


def clear_total_rows(ws, marker='TOTAL'):
    """Clear any existing total rows before placing a new one."""
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and marker in v.upper():
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= r <= mr.max_row:
                    try:
                        ws.unmerge_cells(str(mr))
                    except Exception:
                        pass
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
                ws.cell(row=r, column=c).font = Font()
                ws.cell(row=r, column=c).border = Border()