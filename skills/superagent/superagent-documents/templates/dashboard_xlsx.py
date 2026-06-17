#!/usr/bin/env python3
"""
openpyxl multi-sheet dashboard template — KPI cards + charts + conditional formatting.

REPRODUCE WITH MODIFICATIONS:
1. Replace DATA_PATH with your sample data JSON
2. Adjust CATS_INC / CATS_EXP / GOALS / INVESTMENTS dicts
3. Tweak kpi_card() calls to change card layout
4. Run: python3 dashboard_xlsx.py

REQUIREMENTS:
- pip install openpyxl
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# === CONFIG — EDIT THIS ===
DATA_PATH = "sample_data.json"
OUTPUT_PATH = "laporan_keuangan_v2.xlsx"
CURRENCY_FORMAT = '"Rp "#,##0'

CATS_INC = {"gaji": "Gaji", "bonus": "Bonus", "freelance": "Freelance", "dividen": "Dividen"}
CATS_EXP = {"makanan":"Makanan", "transport":"Transport", "belanja":"Belanja",
            "tagihan":"Tagihan", "hiburan":"Hiburan", "kesehatan":"Kesehatan"}

# === STYLES ===
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='E5E7EB')
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def style_subheader(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER

def kpi_card(ws, row, col_start, col_end, title, value, sublabel, color="2563EB"):
    """Place a KPI card spanning (row, col_start:col_end)."""
    sc, ec = get_column_letter(col_start), get_column_letter(col_end)
    # Title band
    ws.merge_cells(f'{sc}{row}:{ec}{row}')
    cell = ws[f'{sc}{row}']
    cell.value = title
    cell.fill = PatternFill('solid', fgColor=color)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Value
    ws.merge_cells(f'{sc}{row+1}:{ec}{row+1}')
    v = ws[f'{sc}{row+1}']
    v.value = value
    v.font = Font(bold=True, color=color, size=18)
    v.alignment = Alignment(horizontal='center', vertical='center')
    # Sublabel
    ws.merge_cells(f'{sc}{row+2}:{ec}{row+2}')
    s = ws[f'{sc}{row+2}']
    s.value = sublabel
    s.font = Font(italic=True, color='666666', size=9)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row+1].height = 36
    ws.row_dimensions[row+2].height = 18

# === BUILD WORKBOOK ===
def build():
    with open(DATA_PATH) as f:
        data = json.load(f)

    # Normalize kategori to lowercase
    tx = data["transaksi"]
    for t in tx:
        t["kategori"] = t["kategori"].lower()
    # Filter out empty rows (json.loads from google sheets often has trailing empty rows)
    tx = [t for t in tx if t.get("tanggal") or t.get("kategori")]

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard_sheet(wb, tx)
    build_transaksi_sheet(wb, tx)
    build_kategori_sheet(wb, tx)
    build_goals_sheet(wb, data.get("goals", []))
    build_investasi_sheet(wb, data.get("investments", []))
    build_budget_sheet(wb, data.get("categories", {}))
    build_laporan_sheet(wb, tx)

    wb.save(OUTPUT_PATH)
    print(f"✅ Saved: {OUTPUT_PATH} ({Path(OUTPUT_PATH).stat().st_size//1024} KB)")

def build_dashboard_sheet(wb, tx):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    total_in = sum(t["jumlah"] for t in tx if t["tipe"] == "Pemasukan")
    total_out = sum(t["jumlah"] for t in tx if t["tipe"] == "Pengeluaran")
    net = total_in - total_out
    savings_rate = net / total_in if total_in > 0 else 0

    # KPI cards
    kpi_card(ws, 2, 1, 3, "💰 PEMASUKAN",  f"Rp {total_in/1e6:.2f}jt",
             f"{sum(1 for t in tx if t['tipe']=='Pemasukan')} transaksi", "10B981")
    kpi_card(ws, 2, 4, 6, "💸 PENGELUARAN", f"Rp {total_out/1e6:.2f}jt",
             f"{sum(1 for t in tx if t['tipe']=='Pengeluaran')} transaksi", "EF4444")
    kpi_card(ws, 2, 7, 9, "📈 NET", f"Rp {net/1e6:+.2f}jt",
             f"Savings: {savings_rate*100:.1f}%", "3B82F6")

    # Monthly summary table
    ws['A7'] = "📊 RINGKASAN BULANAN"
    ws['A7'].font = Font(bold=True, size=12)
    ws.merge_cells('A7:D7')
    headers = ["Bulan", "Pemasukan", "Pengeluaran", "Net"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=8, column=c, value=h)
    style_subheader(ws, 8, 4)

    by_month = {}
    for t in tx:
        m = t["tanggal"][:7] if isinstance(t["tanggal"], str) else t["tanggal"].strftime("%Y-%m")
        if m not in by_month:
            by_month[m] = {"income": 0, "expense": 0}
        if t["tipe"] == "Pemasukan":
            by_month[m]["income"] += t["jumlah"]
        else:
            by_month[m]["expense"] += t["jumlah"]

    row = 9
    for m in sorted(by_month):
        inc = by_month[m]["income"]
        out = by_month[m]["expense"]
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=inc).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=3, value=out).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = CURRENCY_FORMAT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
        row += 1

    # Bar chart from monthly summary
    if by_month:
        bar = BarChart()
        bar.type = "col"
        bar.style = 11
        bar.title = "Income vs Expense per Bulan"
        bar.y_axis.title = "Jumlah (Rp)"
        bar.x_axis.title = "Bulan"
        data_ref = Reference(ws, min_col=2, min_row=8, max_col=3, max_row=8 + len(by_month))
        cats_ref = Reference(ws, min_col=1, min_row=9, max_row=8 + len(by_month))
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.height = 9
        bar.width = 18
        ws.add_chart(bar, f"A{row + 2}")

def build_transaksi_sheet(wb, tx):
    ws = wb.create_sheet("Transaksi")
    ws.sheet_view.showGridLines = False
    headers = ["Tanggal", "Tipe", "Kategori", "Jumlah", "Akun", "Catatan"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, t in enumerate(tx, start=2):
        ws.cell(row=i, column=1, value=t.get("tanggal"))
        ws.cell(row=i, column=2, value=t.get("tipe"))
        ws.cell(row=i, column=3, value=t.get("kategori"))
        c4 = ws.cell(row=i, column=4, value=t.get("jumlah"))
        c4.number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=t.get("akun", ""))
        ws.cell(row=i, column=6, value=t.get("catatan", ""))
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = BORDER
    # Column widths
    widths = [14, 14, 14, 16, 12, 32]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    # Conditional: highlight large expenses
    if len(tx) > 1:
        ws.conditional_formatting.add(
            f'D2:D{len(tx)+1}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           mid_type='percentile', mid_value=50, mid_color='FED7AA',
                           end_type='max', end_color='EF4444')
        )
    # Total row
    total_row = len(tx) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{len(tx)+1})").number_format = CURRENCY_FORMAT
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.freeze_panes = 'A2'

def build_kategori_sheet(wb, tx):
    ws = wb.create_sheet("Kategori")
    ws.sheet_view.showGridLines = False
    ws['A1'] = "Kategori"; ws['B1'] = "Total"; ws['C1'] = "Jumlah Transaksi"; ws['D1'] = "Rata-rata"
    style_header(ws, 1, 4)
    by_cat = {}
    for t in tx:
        if t["tipe"] == "Pengeluaran":
            k = t["kategori"]
            if k not in by_cat:
                by_cat[k] = {"total": 0, "count": 0}
            by_cat[k]["total"] += t["jumlah"]
            by_cat[k]["count"] += 1
    for i, (k, v) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]["total"]), start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v["total"]).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=v["count"])
        ws.cell(row=i, column=4, value=f"=B{i}/C{i}").number_format = CURRENCY_FORMAT
        for c in range(1, 5):
            ws.cell(row=i, column=c).border = BORDER
    # Pie chart
    if by_cat:
        pie = PieChart()
        pie.title = "Distribusi Pengeluaran"
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(by_cat) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(by_cat) + 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.dataLabels = DataLabelList(showPercent=True)
        pie.height = 10
        pie.width = 16
        ws.add_chart(pie, "F2")
    for c, w in zip([1, 2, 3, 4], [16, 16, 18, 16]):
        ws.column_dimensions[get_column_letter(c)].width = w

def build_goals_sheet(wb, goals):
    ws = wb.create_sheet("Goals")
    ws.sheet_view.showGridLines = False
    headers = ["Goal", "Target", "Current", "Progress", "Sisa", "Deadline"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, g in enumerate(goals, start=2):
        ws.cell(row=i, column=1, value=g["nama"])
        ws.cell(row=i, column=2, value=g["target"]).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=g["current"]).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=f"=C{i}/B{i}").number_format = '0.0%'
        ws.cell(row=i, column=5, value=f"=B{i}-C{i}").number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=6, value=g.get("deadline", ""))
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = BORDER
    # Conditional: 100% green, low red
    if goals:
        ws.conditional_formatting.add(
            f'D2:D{len(goals)+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FCA5A5',
                           mid_type='num', mid_value=0.5, mid_color='FED7AA',
                           end_type='num', end_value=1, end_color='86EFAC')
        )
    for c, w in zip([1, 2, 3, 4, 5, 6], [22, 16, 16, 12, 16, 14]):
        ws.column_dimensions[get_column_letter(c)].width = w

def build_investasi_sheet(wb, invs):
    ws = wb.create_sheet("Investasi")
    ws.sheet_view.showGridLines = False
    headers = ["Nama", "Jenis", "Modal", "Nilai", "Return", "%", "Tanggal Beli"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, inv in enumerate(invs, start=2):
        ws.cell(row=i, column=1, value=inv["nama"])
        ws.cell(row=i, column=2, value=inv["jenis"])
        ws.cell(row=i, column=3, value=inv["modal"]).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=inv["nilai"]).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=f"=D{i}-C{i}").number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=6, value=f"=(D{i}-C{i})/C{i}*100").number_format = '0.00"%"'
        ws.cell(row=i, column=7, value=inv.get("tanggal_beli", ""))
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = BORDER
    # Conditional: positive return green, negative red
    if invs:
        ws.conditional_formatting.add(
            f'F2:F{len(invs)+1}',
            ColorScaleRule(start_type='min', start_color='FCA5A5',
                           mid_type='num', mid_value=0, mid_color='FFFFFF',
                           end_type='max', end_color='86EFAC')
        )
    for c, w in zip([1, 2, 3, 4, 5, 6, 7], [22, 12, 16, 16, 16, 12, 14]):
        ws.column_dimensions[get_column_letter(c)].width = w

def build_budget_sheet(wb, categories):
    """Build a budget plan sheet with monthly limits per category."""
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    headers = ["Kategori", "Budget Bulanan", "Aktual", "Sisa", "% Used"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    # Default budget amounts (customize per user)
    default_budgets = {"makanan": 1500000, "transport": 500000, "belanja": 1000000,
                       "tagihan": 800000, "hiburan": 500000, "kesehatan": 300000}
    row = 2
    for k, v in CATS_EXP.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=default_budgets.get(k, 0)).number_format = CURRENCY_FORMAT
        # Aktual via SUMIF (placeholder — user fills in via actual tx)
        ws.cell(row=row, column=3, value=0).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=5, value=f"=IFERROR(C{row}/B{row},0)").number_format = '0.0%'
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
        row += 1
    if CATS_EXP:
        ws.conditional_formatting.add(
            f'E2:E{len(CATS_EXP)+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='86EFAC',
                           mid_type='num', mid_value=0.75, mid_color='FED7AA',
                           end_type='num', end_value=1, end_color='FCA5A5')
        )
    for c, w in zip([1, 2, 3, 4, 5], [16, 18, 18, 16, 12]):
        ws.column_dimensions[get_column_letter(c)].width = w

def build_laporan_sheet(wb, tx):
    """Build a summary laporan sheet with monthly P/L."""
    ws = wb.create_sheet("Laporan")
    ws.sheet_view.showGridLines = False
    headers = ["Bulan", "Pemasukan", "Pengeluaran", "Net", "Savings Rate"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    by_month = {}
    for t in tx:
        m = t["tanggal"][:7] if isinstance(t["tanggal"], str) else t["tanggal"].strftime("%Y-%m")
        if m not in by_month:
            by_month[m] = {"income": 0, "expense": 0}
        if t["tipe"] == "Pemasukan":
            by_month[m]["income"] += t["jumlah"]
        else:
            by_month[m]["expense"] += t["jumlah"]
    for i, m in enumerate(sorted(by_month), start=2):
        inc = by_month[m]["income"]
        out = by_month[m]["expense"]
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2, value=inc).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=out).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=f"=IFERROR(D{i}/B{i},0)").number_format = '0.0%'
        for c in range(1, 6):
            ws.cell(row=i, column=c).border = BORDER
    # Line chart
    if by_month:
        line = LineChart()
        line.title = "Trend Net Bulanan"
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(by_month)+1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(by_month)+1)
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        line.height = 9
        line.width = 18
        ws.add_chart(line, "G2")
    for c, w in zip([1, 2, 3, 4, 5], [12, 16, 16, 16, 14]):
        ws.column_dimensions[get_column_letter(c)].width = w

if __name__ == "__main__":
    build()
